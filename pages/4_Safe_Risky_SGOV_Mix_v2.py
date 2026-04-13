"""
Safe/Risky/Cash Mix Strategy — 파라미터 최적화 스크립트
실행: python optimize_strategy.py
결과: optimization_results_YYYYMMDD.xlsx (동일 폴더에 저장)

필요 라이브러리: pip install yfinance pandas numpy openpyxl
"""

import yfinance as yf
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from itertools import product
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# ══════════════════════════════════════════════════════════════════════════════
# 1. 설정 (필요 시 수정)
# ══════════════════════════════════════════════════════════════════════════════
TICKER_SAFE  = "SPY"
TICKER_RISKY = "UPRO"
TICKER_CASH  = "SGOV"
TICKER_RATE  = "^TNX"
START_DATE   = "2020-01-01"
INITIAL_CAP  = 100_000_000
FEE_RATE     = 0.0002  # 0.02%

# ── 탐색 파라미터 그리드 ──────────────────────────────────────────────────────
PARAM_GRID = {
    "ma_entry"      : [60, 80, 100, 120, 150, 200],   # Bull 진입 이평선
    "ma_exit_ratio" : [0.4, 0.5, 0.6, 0.75, 1.0],    # 퇴출 MA = 진입 MA × 비율 (1.0 = 비대칭 미사용)
    "rate_ma"       : [60, 90, 120, 150],              # 금리 이평선
    "confirm_days"  : [1, 3, 5],                       # Whipsaw 필터 확정일
    "exposure_ratio": [0.4, 0.5, 0.6, 0.7, 0.8],     # Bull_Mix 공격 비중
    "rebal_thresh"  : [0.03, 0.05, 0.08],             # 리밸런싱 임계값
}
# 총 조합 수 = 6×5×4×3×5×3 = 5,400개

# ══════════════════════════════════════════════════════════════════════════════
# 2. 데이터 로드
# ══════════════════════════════════════════════════════════════════════════════
def load_data():
    print("📡 데이터 다운로드 중...")
    tickers = [TICKER_SAFE, TICKER_RISKY, TICKER_CASH, TICKER_RATE]
    raw = yf.download(tickers, start="2000-01-01", progress=False, auto_adjust=True)
    df  = raw['Close'].ffill()
    df  = df.loc[~df.index.duplicated(keep='first')].sort_index()
    print(f"  ✔ 다운로드 완료: {df.index[0].date()} ~ {df.index[-1].date()}, {len(df)}거래일\n")
    return df

# ══════════════════════════════════════════════════════════════════════════════
# 3. 단일 백테스트 함수
# ══════════════════════════════════════════════════════════════════════════════
def run_backtest(df_raw, returns_df,
                 ma_entry, ma_exit, rate_ma,
                 confirm_days, exposure_ratio, rebal_thresh,
                 use_rate_filter=True):

    sim_start = pd.to_datetime(START_DATE)
    series_safe  = df_raw[TICKER_SAFE]
    series_rate  = df_raw[TICKER_RATE]
    cash_avail   = TICKER_CASH in df_raw.columns

    # 지표 계산
    ma_safe_entry = series_safe.rolling(int(ma_entry)).mean()
    ma_safe_exit_ = series_safe.rolling(int(ma_exit)).mean()
    ma_rate_s     = series_rate.rolling(int(rate_ma)).mean()
    spy_vol       = series_safe.pct_change().rolling(20).std() * np.sqrt(252)

    # 비대칭 Bull/Bear 상태 머신
    raw_bull_entry = series_safe > ma_safe_entry
    raw_bear_exit  = series_safe < ma_safe_exit_
    state_arr = np.zeros(len(series_safe), dtype=int)
    for i in range(1, len(series_safe)):
        if state_arr[i-1] == 0:
            state_arr[i] = 1 if raw_bull_entry.iloc[i] else 0
        else:
            state_arr[i] = 0 if raw_bear_exit.iloc[i] else 1
    raw_bull = pd.Series(state_arr == 1, index=df_raw.index)

    # Whipsaw 필터
    cd = int(confirm_days)
    is_bull_confirmed = raw_bull.rolling(cd).min().fillna(0).astype(bool)
    is_bear_confirmed = (~raw_bull).rolling(cd).min().fillna(0).astype(bool)
    bull_signal = pd.Series(np.nan, index=df_raw.index, dtype='float64')
    bull_signal = bull_signal.where(~is_bull_confirmed, other=1.0)
    bull_signal = bull_signal.where(~is_bear_confirmed, other=0.0)
    is_bull = bull_signal.ffill().fillna(0.0).astype(bool)

    is_hike = series_rate > ma_rate_s

    # 상태 결정 + T+1 지연
    conditions = [
        ~is_bull,
        is_bull & (~is_hike | ~use_rate_filter),
        is_bull & is_hike & use_rate_filter,
    ]
    raw_state   = pd.Series(np.select(conditions, ["Bear","Bull_Full","Bull_Mix"], default="Bear"), index=df_raw.index)
    trade_state = raw_state.shift(1)

    # 시뮬레이션 범위
    df_sim      = df_raw.loc[sim_start:].copy()
    trade_state = trade_state.loc[sim_start:].fillna("Bear")
    returns_sim = returns_df.loc[sim_start:]
    spy_vol_sim = spy_vol.loc[sim_start:]

    def get_weights(state, today_vol):
        # Bull_Full: 변동성 사이징 미적용 (CAGR 보호)
        # Bull_Mix: exposure_ratio 기준 배분
        if state == "Bear":
            return {TICKER_SAFE: 1.0}
        elif state == "Bull_Full":
            return {TICKER_RISKY: 1.0}
        elif state == "Bull_Mix":
            rw = exposure_ratio
            cw = 1.0 - rw
            w  = {TICKER_RISKY: rw}
            if cw > 1e-6 and cash_avail:
                w[TICKER_CASH] = cw
            return w
        return {TICKER_SAFE: 1.0}

    equity = float(INITIAL_CAP)
    peak   = equity
    equity -= equity * FEE_RATE

    curr_w     = get_weights(trade_state.iloc[0], spy_vol_sim.iloc[0])
    daily_rets = []
    dds        = []
    n_trades   = 0

    for i in range(len(df_sim)):
        today    = df_sim.index[i]
        state    = trade_state.iloc[i]
        vol_now  = spy_vol_sim.iloc[i] if i < len(spy_vol_sim) else np.nan
        target_w = {k: v for k, v in get_weights(state, vol_now).items() if v > 0}

        day_ret = 0.0
        if i > 0:
            for tk, w in curr_w.items():
                if tk in returns_sim.columns:
                    r = returns_sim.loc[today, tk]
                    day_ret += w * (0.0 if pd.isna(r) else float(r))

        equity *= (1.0 + day_ret)
        daily_rets.append(day_ret)

        # 리밸런싱: 상태 변경 OR 비중 변화 > 임계값
        state_changed  = (curr_w.keys() != target_w.keys())
        weight_changed = any(
            abs(curr_w.get(k, 0) - target_w.get(k, 0)) >= rebal_thresh
            for k in set(list(curr_w.keys()) + list(target_w.keys()))
        )
        if state_changed or weight_changed:
            equity -= equity * FEE_RATE
            curr_w  = target_w
            n_trades += 1

        if equity > peak: peak = equity
        dds.append((equity - peak) / peak)

    # 성과 계산
    days  = (df_sim.index[-1] - df_sim.index[0]).days
    cagr  = (equity / INITIAL_CAP) ** (365.0 / days) - 1 if days > 0 else 0
    mdd   = min(dds)
    ret_s = pd.Series(daily_rets)
    sharpe = float(((ret_s.mean() - 0.05/252) / ret_s.std()) * np.sqrt(252)) if ret_s.std() > 0 else 0
    calmar = abs(cagr / mdd) if mdd != 0 else 0

    return {
        "CAGR(%)":    round(cagr * 100, 2),
        "MDD(%)":     round(mdd * 100,  2),
        "Sharpe":     round(sharpe,      3),
        "Calmar":     round(calmar,      3),
        "Trades":     n_trades,
        "Final(억)":  round(equity / 1e8, 2),
    }

# ══════════════════════════════════════════════════════════════════════════════
# 4. 그리드 서치
# ══════════════════════════════════════════════════════════════════════════════
def grid_search(df_raw):
    returns_df = df_raw.pct_change()
    results    = []
    keys       = list(PARAM_GRID.keys())
    combos     = list(product(*[PARAM_GRID[k] for k in keys]))
    total      = len(combos)

    print(f"🔍 총 {total:,}개 파라미터 조합 탐색 시작...\n")

    for idx, combo in enumerate(combos):
        params = dict(zip(keys, combo))

        ma_entry  = params["ma_entry"]
        ma_exit   = max(5, int(ma_entry * params["ma_exit_ratio"]))
        rate_ma   = params["rate_ma"]
        conf_days = params["confirm_days"]
        exp_ratio = params["exposure_ratio"]
        rebal_thr = params["rebal_thresh"]

        try:
            perf = run_backtest(
                df_raw, returns_df,
                ma_entry=ma_entry, ma_exit=ma_exit,
                rate_ma=rate_ma, confirm_days=conf_days,
                exposure_ratio=exp_ratio, rebal_thresh=rebal_thr
            )
        except Exception as e:
            continue

        results.append({
            "ma_entry":      ma_entry,
            "ma_exit":       ma_exit,
            "ma_exit_ratio": params["ma_exit_ratio"],
            "rate_ma":       rate_ma,
            "confirm_days":  conf_days,
            "exposure_ratio":exp_ratio,
            "rebal_thresh(%)": round(rebal_thr * 100, 1),
            **perf
        })

        if (idx + 1) % 500 == 0 or (idx + 1) == total:
            done = idx + 1
            pct  = done / total * 100
            best = max(results, key=lambda x: x["Sharpe"])
            print(f"  [{done:>5}/{total}] {pct:.1f}% 완료 | 현재 최고 Sharpe: {best['Sharpe']} "
                  f"(CAGR {best['CAGR(%)']:+.1f}%, MDD {best['MDD(%)']:.1f}%)")

    return pd.DataFrame(results)

# ══════════════════════════════════════════════════════════════════════════════
# 5. 결과 저장
# ══════════════════════════════════════════════════════════════════════════════
def save_results(df_res):
    df_res = df_res.sort_values("Sharpe", ascending=False).reset_index(drop=True)

    filename = f"optimization_results_{datetime.today().strftime('%Y%m%d')}.xlsx"
    wb = openpyxl.Workbook()

    # ── Sheet 1: 전체 결과 (Sharpe 기준 정렬) ────────────────────────────────
    ws1 = wb.active
    ws1.title = "전체결과_Sharpe순"
    _write_sheet(ws1, df_res, "📊 전체 결과 (Sharpe 내림차순)")

    # ── Sheet 2: CAGR Top 50 ────────────────────────────────────────────────
    ws2 = wb.create_sheet("CAGR_Top50")
    top_cagr = df_res.nlargest(50, "CAGR(%)")
    _write_sheet(ws2, top_cagr.reset_index(drop=True), "🚀 CAGR Top 50")

    # ── Sheet 3: MDD 최소 Top 50 ────────────────────────────────────────────
    ws3 = wb.create_sheet("MDD_Best50")
    top_mdd = df_res.nlargest(50, "MDD(%)")  # MDD는 음수이므로 nlargest = 낙폭 가장 작은 것
    _write_sheet(ws3, top_mdd.reset_index(drop=True), "🛡️ MDD Best 50 (낙폭 최소)")

    # ── Sheet 4: Calmar Top 50 ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Calmar_Top50")
    top_cal = df_res.nlargest(50, "Calmar")
    _write_sheet(ws4, top_cal.reset_index(drop=True), "⚖️ Calmar Top 50 (CAGR/MDD 균형)")

    # ── Sheet 5: 베스트 요약 ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("🏆 베스트 요약")
    _write_summary(ws5, df_res)

    wb.save(filename)
    return filename, df_res

def _write_sheet(ws, df, title):
    GREEN = "00703C"
    h_fill  = PatternFill("solid", start_color=GREEN, end_color=GREEN)
    h_font  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    d_font  = Font(name="맑은 고딕", size=10)
    alt_fill= PatternFill("solid", start_color="F2FAF5", end_color="F2FAF5")
    ctr     = Alignment(horizontal="center", vertical="center")
    rgt     = Alignment(horizontal="right",  vertical="center")

    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    ws["A1"].value     = title
    ws["A1"].font      = Font(name="맑은 고딕", bold=True, size=13, color=GREEN)
    ws["A1"].alignment = ctr
    ws.row_dimensions[1].height = 26

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.font  = h_font
        cell.fill  = h_fill
        cell.alignment = ctr
        ws.column_dimensions[get_column_letter(ci)].width = max(len(str(col)) + 4, 12)

    for ri, row in df.iterrows():
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri + 3, column=ci, value=val)
            cell.font = d_font
            if fill: cell.fill = fill
            cell.alignment = rgt if ci > 3 else ctr
        ws.row_dimensions[ri + 3].height = 16

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(df.columns))}{len(df) + 2}"

def _write_summary(ws, df_res):
    GREEN = "00703C"
    t_font = Font(name="맑은 고딕", bold=True, size=14, color=GREEN)
    h_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    d_font = Font(name="맑은 고딕", size=11)
    h_fill = PatternFill("solid", start_color=GREEN, end_color=GREEN)
    ctr    = Alignment(horizontal="center", vertical="center")

    ws["A1"].value = "🏆 파라미터 최적화 베스트 결과 요약"
    ws["A1"].font  = t_font
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = ctr
    ws.row_dimensions[1].height = 30

    sections = [
        ("🥇 Sharpe 최고 (위험 대비 수익 최적)",  df_res.nlargest(1, "Sharpe").iloc[0]),
        ("🚀 CAGR 최고 (수익률 최우선)",           df_res.nlargest(1, "CAGR(%)").iloc[0]),
        ("🛡️ MDD 최소 (안정성 최우선)",            df_res.nlargest(1, "MDD(%)").iloc[0]),
        ("⚖️ Calmar 최고 (수익/위험 균형)",        df_res.nlargest(1, "Calmar").iloc[0]),
    ]
    # Sharpe≥1 AND MDD≥-35% 필터
    balanced = df_res[(df_res["Sharpe"] >= 0.8) & (df_res["MDD(%)"] >= -40)]
    if len(balanced) > 0:
        sections.append(("💎 균형 추천 (Sharpe≥0.8 & MDD≥-40%)", balanced.nlargest(1, "CAGR(%)").iloc[0]))

    row = 3
    for title, best in sections:
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"].value = title
        ws[f"A{row}"].font  = h_font
        ws[f"A{row}"].fill  = h_fill
        ws[f"A{row}"].alignment = ctr
        ws.row_dimensions[row].height = 22
        row += 1

        params_display = [
            ("Bull 진입 MA",    f"{int(best['ma_entry'])}일"),
            ("Bear 퇴출 MA",    f"{int(best['ma_exit'])}일"),
            ("금리 MA",         f"{int(best['rate_ma'])}일"),
            ("Whipsaw 확정",    f"{int(best['confirm_days'])}일"),
            ("Bull_Mix 비중",   f"{best['exposure_ratio']*100:.0f}%"),
            ("리밸런싱 임계",   f"{best['rebal_thresh(%)']}%"),
            ("CAGR",            f"{best['CAGR(%)']}%"),
            ("MDD",             f"{best['MDD(%)']}%"),
            ("Sharpe",          f"{best['Sharpe']}"),
            ("Calmar",          f"{best['Calmar']}"),
            ("매매횟수",        f"{int(best['Trades'])}회"),
            ("최종잔고",        f"{best['Final(억)']}억원"),
        ]

        col = 1
        for i, (k, v) in enumerate(params_display):
            r = row + (i // 6)
            c = col + (i % 6)
            label_cell = ws.cell(row=r, column=c, value=k)
            label_cell.font = Font(name="맑은 고딕", bold=True, size=10, color="555555")
            label_cell.alignment = ctr
            ws.row_dimensions[r].height = 20
            ws.column_dimensions[get_column_letter(c)].width = 16

            val_cell = ws.cell(row=r + 1, column=c, value=v)
            val_cell.font = Font(name="맑은 고딕", bold=True, size=12,
                                  color="C0392B" if "MDD" in k else "1A5276" if "CAGR" in k else "000000")
            val_cell.alignment = ctr
            ws.row_dimensions[r + 1].height = 22

        row += 4

# ══════════════════════════════════════════════════════════════════════════════
# 6. 메인 실행
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 60)
    print("  Safe/Risky/Cash Mix Strategy — 파라미터 최적화")
    print(f"  기간: {START_DATE} ~ 현재")
    print(f"  초기자본: {INITIAL_CAP:,}원 / 수수료: {FEE_RATE*100:.2f}%")
    print("=" * 60 + "\n")

    df_raw  = load_data()
    df_res  = grid_search(df_raw)

    print(f"\n✅ 탐색 완료! 유효 결과: {len(df_res):,}개\n")

    # 상위 결과 콘솔 출력
    print("━" * 60)
    print("📊 Sharpe 기준 TOP 10:")
    top10 = df_res.nlargest(10, "Sharpe")[
        ["ma_entry","ma_exit","rate_ma","confirm_days","exposure_ratio",
         "rebal_thresh(%)","CAGR(%)","MDD(%)","Sharpe","Calmar","Trades"]
    ]
    print(top10.to_string(index=False))

    print("\n💎 균형 추천 (Sharpe≥0.8 & MDD≥-40%):")
    balanced = df_res[(df_res["Sharpe"] >= 0.8) & (df_res["MDD(%)"] >= -40)]
    if len(balanced) > 0:
        best_bal = balanced.nlargest(5, "CAGR(%)")
        print(best_bal[["ma_entry","ma_exit","rate_ma","confirm_days",
                         "exposure_ratio","CAGR(%)","MDD(%)","Sharpe","Calmar"]].to_string(index=False))
    else:
        print("  해당 조건을 만족하는 조합이 없습니다. MDD 기준을 완화해 보세요.")

    filename, _ = save_results(df_res)
    print(f"\n📥 결과 저장 완료: {filename}")
    print("   → Excel 파일에서 시트별 분석 확인하세요!")